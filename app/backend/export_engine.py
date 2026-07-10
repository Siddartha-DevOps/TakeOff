"""
TakeOff.ai — Rich export engine: grouped/filtered Excel, CSV, and PDF
quantity reports across one or more drawings.

Closes memory/TOGAL_PARITY_REAUDIT.md #14 (Phase D): "Rich export — Excel +
PDF, with grouping (3 levels), filtering, drawing selection, export
multiplier, inline editable grid." Also closes the "project export = first
drawing only" bug it names — routes/export_routes.py's old export_project()
queried every drawing in a project, then kept only `drawings[0]`. Every
function here processes whatever set of drawings/rows it's given, in full.

Data source: TakeoffResult.quantities_data (the AI's trade/item/quantity/
unit breakdown per drawing) — not Condition/Detection.condition_id.
Condition-based cost (quantity * unit_cost * (1+waste%), per
frontend/src/pages/Takeoff.jsx's conditionCostTotals) is a live,
client-side-only concept today: assigning a shape to a Condition
(assignSelectionToCondition) calls annotationStore directly and never hits
the backend, so Detection.condition_id is never populated by the real app
flow — confirmed by reading the assignment code path, not assumed. A
backend-side join through it would silently show empty/incomplete data for
every real project. quantities_data is the one dataset that's genuinely
persisted per-drawing and safe to aggregate project-wide; using it instead
of Condition is a deliberate, documented scope decision.

The "inline editable grid" requirement is satisfied by construction:
generate_report() below renders exactly the rows it's handed, verbatim —
routes/export_routes.py's preview endpoint returns rows in this same
shape, the frontend lets a user edit quantities / exclude rows, and
whatever it submits to the generate endpoint is what ends up in the file.
No re-querying happens in between, so there's no way for the export to
silently diverge from what was previewed and edited.

No per-row cost/subtotal: rows can mix units (sf/lf/ea) even within one
trade, so a naive sum would be meaningless. Each row carries its own
quantity+unit; totalling is left to whatever spreadsheet/estimating tool
consumes the export, matching Togal's own "quantities hand off to
estimating partners" positioning (memory/TOGAL_PARITY_REAUDIT.md §1.8).
"""

import csv
import io
import json
from typing import Optional

import models

GROUP_DIMENSIONS = ("drawing", "trade", "item")
_GROUP_FIELD = {"drawing": "drawing_name", "trade": "trade", "item": "item"}


def extract_rows(db, drawing: "models.Drawing") -> list[dict]:
    """
    One row per quantities_data line item from the given Drawing's latest
    TakeoffResult. A drawing with no result yet (still processing)
    contributes zero rows, not an error — a project export explicitly
    spans drawings that may be at different stages.
    """
    result = db.query(models.TakeoffResult).filter(
        models.TakeoffResult.drawing_id == drawing.id
    ).order_by(models.TakeoffResult.created_at.desc()).first()
    if not result or not result.quantities_data:
        return []

    try:
        quantities = json.loads(result.quantities_data)
    except (json.JSONDecodeError, TypeError):
        return []
    if not isinstance(quantities, list):
        return []

    drawing_name = drawing.sheet_number or drawing.sheet_name or drawing.original_filename
    rows = []
    for i, item in enumerate(quantities):
        if not isinstance(item, dict):
            continue
        try:
            quantity = float(item.get("quantity", 0) or 0)
        except (TypeError, ValueError):
            quantity = 0.0
        rows.append({
            "row_id": f"{drawing.id}:{i}",
            "drawing_id": drawing.id,
            "drawing_name": drawing_name,
            "trade": item.get("trade") or "Uncategorized",
            "item": item.get("item") or "Untitled",
            "quantity": quantity,
            "unit": item.get("unit") or "",
        })
    return rows


def filter_rows(rows: list[dict], drawing_ids: Optional[list[int]] = None, trades: Optional[list[str]] = None) -> list[dict]:
    out = rows
    if drawing_ids:
        wanted_ids = set(drawing_ids)
        out = [r for r in out if r["drawing_id"] in wanted_ids]
    if trades:
        wanted_trades = set(trades)
        out = [r for r in out if r["trade"] in wanted_trades]
    return out


def apply_multiplier(rows: list[dict], multiplier: float) -> list[dict]:
    """
    "An export multiplier (per floor/building)" per
    memory/TOGAL_PARITY_REAUDIT.md §1.8 — e.g. one measured unit repeated
    across N identical floors. multiplier=1.0 is a no-op passthrough
    (returns the same list, not a copy) so callers can apply it
    unconditionally without a branch.
    """
    if multiplier == 1.0:
        return rows
    return [{**r, "quantity": round(r["quantity"] * multiplier, 4)} for r in rows]


def build_grouped_sections(rows: list[dict], group_by: Optional[list[str]] = None) -> dict:
    """
    Nests `rows` by up to 3 dimensions from GROUP_DIMENSIONS (in the given
    order — e.g. ["trade", "drawing"] groups by trade first, then by sheet
    within each trade). Returns a tree:
      {"label": str | None, "rows": [...] (only on leaf nodes), "children": [...]}
    An empty/None group_by yields a single root node with `rows` set
    directly — flat table, no sections.
    """
    dims = [d for d in (group_by or []) if d in GROUP_DIMENSIONS][:3]

    def _build(subset: list[dict], remaining: list[str]) -> dict:
        if not remaining:
            return {"label": None, "rows": subset, "children": []}
        dim = remaining[0]
        field = _GROUP_FIELD[dim]
        order: list[str] = []
        buckets: dict[str, list[dict]] = {}
        for r in subset:
            key = r.get(field) or "—"
            if key not in buckets:
                buckets[key] = []
                order.append(key)
            buckets[key].append(r)
        children = []
        for key in order:
            child = _build(buckets[key], remaining[1:])
            child["label"] = key
            children.append(child)
        return {"label": None, "rows": [], "children": children}

    return _build(rows, dims)


def render_excel(sections: dict, title: str = "Takeoff Export") -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Export"

    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    group_fonts = [Font(bold=True, size=14, color="1F2937"), Font(bold=True, size=12, color="334155"), Font(bold=True, size=11, color="475569")]
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=16)
    row_idx = 3
    col_headers = ["Item", "Trade", "Sheet", "Quantity", "Unit"]

    def write_leaf_table(rows):
        nonlocal row_idx
        for c, h in enumerate(col_headers, start=1):
            cell = ws.cell(row=row_idx, column=c, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        row_idx += 1
        for r in rows:
            ws.cell(row=row_idx, column=1, value=r["item"]).border = border
            ws.cell(row=row_idx, column=2, value=r["trade"]).border = border
            ws.cell(row=row_idx, column=3, value=r["drawing_name"]).border = border
            ws.cell(row=row_idx, column=4, value=r["quantity"]).border = border
            ws.cell(row=row_idx, column=5, value=r["unit"]).border = border
            row_idx += 1
        row_idx += 1

    def walk(node, depth):
        nonlocal row_idx
        if node["label"] is not None:
            cell = ws.cell(row=row_idx, column=1, value=node["label"])
            cell.font = group_fonts[min(depth - 1, 2)]
            row_idx += 1
        if node["children"]:
            for child in node["children"]:
                walk(child, depth + 1)
        elif node["rows"]:
            write_leaf_table(node["rows"])

    walk(sections, 0)
    if not sections["children"] and not sections["rows"]:
        ws.cell(row=row_idx, column=1, value="No rows in this export.")

    for col, width in zip("ABCDE", [32, 18, 22, 12, 10]):
        ws.column_dimensions[col].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def render_pdf(sections: dict, title: str = "Takeoff Export") -> io.BytesIO:
    from xml.sax.saxutils import escape

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    styles = getSampleStyleSheet()
    # reportlab's Paragraph interprets a small HTML-like markup subset in its
    # text — group labels can come from rows a user edited in the frontend's
    # inline grid (routes/export_routes.py's /generate accepts arbitrary
    # client-submitted rows), so they're untrusted input by the time they
    # get here. Table cells below are plain strings, not Paragraphs, and
    # aren't markup-parsed — only these Paragraph() calls need escaping.
    story = [Paragraph(escape(title), styles["Title"]), Spacer(1, 0.2 * inch)]
    group_styles = [styles["Heading2"], styles["Heading3"], styles["Heading4"]]

    def leaf_table(rows):
        data = [["Item", "Trade", "Sheet", "Quantity", "Unit"]]
        for r in rows:
            data.append([r["item"], r["trade"], r["drawing_name"], f"{r['quantity']:,.2f}", r["unit"]])
        table = Table(data, hAlign="LEFT", colWidths=[2.3 * inch, 1.2 * inch, 1.3 * inch, 0.9 * inch, 0.6 * inch])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
            ("ALIGN", (3, 1), (3, -1), "RIGHT"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ]))
        return table

    def walk(node, depth):
        if node["label"] is not None:
            style = group_styles[min(depth - 1, 2)]
            story.append(Paragraph(escape(str(node["label"])), style))
        if node["children"]:
            for child in node["children"]:
                walk(child, depth + 1)
        elif node["rows"]:
            story.append(leaf_table(node["rows"]))
            story.append(Spacer(1, 0.15 * inch))

    walk(sections, 0)
    if not sections["children"] and not sections["rows"]:
        story.append(Paragraph("No rows in this export.", styles["Normal"]))

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=letter, title=title,
                             leftMargin=0.6 * inch, rightMargin=0.6 * inch,
                             topMargin=0.6 * inch, bottomMargin=0.6 * inch)
    doc.build(story)
    output.seek(0)
    return output


def render_csv(rows: list[dict]) -> io.BytesIO:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Item", "Trade", "Sheet", "Quantity", "Unit"])
    for r in rows:
        writer.writerow([r["item"], r["trade"], r["drawing_name"], r["quantity"], r["unit"]])
    return io.BytesIO(output.getvalue().encode("utf-8"))


def generate_report(rows: list[dict], fmt: str, group_by: Optional[list[str]] = None, title: str = "Takeoff Export") -> io.BytesIO:
    if fmt == "csv":
        return render_csv(rows)
    sections = build_grouped_sections(rows, group_by)
    if fmt == "excel":
        return render_excel(sections, title=title)
    if fmt == "pdf":
        return render_pdf(sections, title=title)
    raise ValueError(f"Unsupported export format: {fmt}")
