"""
Render a full India estimate (see ``estimate.full_estimate``) to Excel or PDF.

The BOQ is the tender deliverable in India, so it must export cleanly to the two
formats estimators actually submit: a formatted Excel workbook (line items +
abstract + GST summary) and a printable PDF. Both take the dict produced by
``full_estimate`` and return raw bytes for a StreamingResponse.

openpyxl and reportlab are already project dependencies.
"""

from __future__ import annotations

import io


def _money(x) -> str:
    return f"{float(x):,.2f}"


def _summary_rows(estimate: dict) -> list[tuple[str, float]]:
    s = estimate["summary"]
    gst = s["gst"]
    rows = [
        ("Subtotal (pre-tax)", s["subtotal"]),
        (f"Overheads & profit ({s['overhead_profit_pct']:g}%)", s["overhead_profit"]),
        (f"Contingency ({s['contingency_pct']:g}%)", s["contingency"]),
        ("Taxable value", s["taxable_value"]),
    ]
    if gst.get("igst"):
        rows.append((f"IGST ({gst['rate'] * 100:g}%)", gst["igst"]))
    else:
        rows.append((f"CGST ({gst['rate'] * 50:g}%)", gst["cgst"]))
        rows.append((f"SGST ({gst['rate'] * 50:g}%)", gst["sgst"]))
    rows.append(("Grand total", s["grand_total"]))
    return rows


def boq_to_excel(estimate: dict) -> bytes:
    """Formatted BOQ workbook: line items + chapter abstract + tender summary."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ"

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")
    right = Alignment(horizontal="right")

    ws["A1"] = "Bill of Quantities"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Rate book: {estimate.get('edition', '')}"
    ws["A3"] = estimate.get("rate_book_note", "")

    # Line-item header.
    headers = ["Code", "Description", "Unit", "Qty", "Rate (INR)", "Amount (INR)"]
    r = 5
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
    for item in estimate.get("boq", []):
        r += 1
        ws.cell(row=r, column=1, value=item["code"])
        ws.cell(row=r, column=2, value=item["description"])
        ws.cell(row=r, column=3, value=item["unit"])
        ws.cell(row=r, column=4, value=item["quantity"]).alignment = right
        ws.cell(row=r, column=5, value=item["rate"]).alignment = right
        ws.cell(row=r, column=6, value=item["amount"]).alignment = right

    # Chapter abstract.
    r += 2
    ws.cell(row=r, column=1, value="Abstract (by chapter)").font = bold
    for ch in estimate.get("abstract", {}).get("chapters", []):
        r += 1
        ws.cell(row=r, column=2, value=ch["chapter"])
        ws.cell(row=r, column=6, value=ch["amount"]).alignment = right

    # Tender summary.
    r += 2
    ws.cell(row=r, column=1, value="Tender summary (INR)").font = bold
    for label, amount in _summary_rows(estimate):
        r += 1
        ws.cell(row=r, column=2, value=label)
        cell = ws.cell(row=r, column=6, value=round(float(amount), 2))
        cell.alignment = right
        if label == "Grand total":
            cell.font = bold
            ws.cell(row=r, column=2).font = bold

    for col, width in zip("ABCDEF", (12, 46, 8, 12, 14, 16)):
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def boq_to_pdf(estimate: dict) -> bytes:
    """Printable BOQ PDF: line items + abstract + tender summary."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    )
    from reportlab.lib.styles import getSampleStyleSheet

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, title="Bill of Quantities",
                            leftMargin=15 * mm, rightMargin=15 * mm,
                            topMargin=15 * mm, bottomMargin=15 * mm)
    styles = getSampleStyleSheet()
    elems = [Paragraph("Bill of Quantities", styles["Title"])]
    note = estimate.get("rate_book_note")
    if note:
        elems.append(Paragraph(note, styles["Italic"]))
    elems.append(Spacer(1, 6 * mm))

    data = [["Code", "Description", "Unit", "Qty", "Rate", "Amount"]]
    for it in estimate.get("boq", []):
        data.append([it["code"], it["description"], it["unit"],
                     _money(it["quantity"]), _money(it["rate"]), _money(it["amount"])])
    if len(data) == 1:
        data.append(["—", "No priced items (run AUTODETECT first)", "", "", "", ""])
    table = Table(data, colWidths=[22 * mm, 70 * mm, 14 * mm, 20 * mm, 22 * mm, 26 * mm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F5F9")]),
    ]))
    elems.append(table)
    elems.append(Spacer(1, 6 * mm))

    summ = [["Tender summary (INR)", ""]]
    summ += [[label, _money(amount)] for label, amount in _summary_rows(estimate)]
    stable = Table(summ, colWidths=[100 * mm, 40 * mm])
    stable.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("FONTNAME", (0, 0), (0, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("LINEABOVE", (0, -1), (-1, -1), 0.8, colors.black),
    ]))
    elems.append(stable)

    doc.build(elems)
    return buf.getvalue()
