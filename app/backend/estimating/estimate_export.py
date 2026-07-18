"""
Excel export for a saved assemblies estimate.

Renders the estimate snapshot (``{drivers, line_items, by_trade, total}``) into a
formatted .xlsx: line items, a per-trade summary, and the grand total. openpyxl
is a project dependency (also used by estimating/export.py for the India BOQ).
The workbook build is deterministic and unit-tested.
"""

from __future__ import annotations

from typing import Optional


def estimate_to_excel(estimate: dict, *, title: str = "Estimate") -> bytes:
    """Build an .xlsx (bytes) from an assemblies estimate snapshot."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Estimate"

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")
    right = Alignment(horizontal="right")

    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    # Drivers row (what the estimate was measured from)
    drivers = estimate.get("drivers") or {}
    if drivers:
        ws.append(["Measured drivers"])
        ws.cell(row=ws.max_row, column=1).font = bold
        for k, v in drivers.items():
            ws.append([k.replace("_", " "), v])
        ws.append([])

    # Line items
    headers = ["Item", "Trade", "Quantity", "Unit", "Unit cost", "Amount"]
    ws.append(headers)
    hdr_row = ws.max_row
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=hdr_row, column=col)
        c.fill, c.font = header_fill, header_font

    for ln in estimate.get("line_items") or []:
        ws.append([
            ln.get("item"), ln.get("trade"), ln.get("quantity"),
            ln.get("unit"), ln.get("unit_cost"), ln.get("amount"),
        ])
        for col in (3, 5, 6):
            ws.cell(row=ws.max_row, column=col).alignment = right

    ws.append([])

    # Per-trade summary
    by_trade = estimate.get("by_trade") or {}
    if by_trade:
        ws.append(["By trade", "", "", "", "", "Amount"])
        ws.cell(row=ws.max_row, column=1).font = bold
        for trade, amt in by_trade.items():
            ws.append([trade, "", "", "", "", amt])
            ws.cell(row=ws.max_row, column=6).alignment = right

    ws.append(["TOTAL", "", "", "", "", estimate.get("total", 0)])
    total_row = ws.max_row
    for col in (1, 6):
        ws.cell(row=total_row, column=col).font = bold
    ws.cell(row=total_row, column=6).alignment = right

    widths = [34, 18, 12, 8, 12, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
