from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from typing import List, Optional
import schemas
import models
import export_engine
from auth import get_current_user
from database import get_db
import json
import io
import csv
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

router = APIRouter(prefix="/export", tags=["Export"])

# ── Rich export: grouping, filtering, drawing selection, multiplier,
# inline editable grid — memory/TOGAL_PARITY_REAUDIT.md #14 ────────────────
# The two endpoints above are untouched (still the quick single-drawing/
# first-drawing-only Excel+CSV path some part of the UI may already call).
# These are new, additive: a real multi-drawing, multi-format export built
# on export_engine.py. See that module's docstring for why it reads
# TakeoffResult.quantities_data rather than Condition/Detection.condition_id.

def _get_project_for_export(project_id: int, current_user: models.User, db: Session) -> models.Project:
    project = db.query(models.Project).filter(
        models.Project.id == project_id,
        models.Project.organization_id == current_user.organization_id,
    ).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return project


@router.get("/projects/{project_id}/preview")
async def preview_project_export(
    project_id: int,
    drawing_ids: Optional[str] = None,  # comma-separated Drawing ids; omit for all
    trades: Optional[str] = None,       # comma-separated trade names; omit for all
    multiplier: float = 1.0,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    One row per AI-detected quantity line item across every selected
    drawing — fixes the old export_project()'s "first drawing only" bug by
    construction, since every matching Drawing contributes its rows here,
    not just drawings[0]. Returns the same row shape the frontend's inline
    editable grid renders and, after edits, posts to /generate verbatim.
    """
    _get_project_for_export(project_id, current_user, db)

    drawings_query = db.query(models.Drawing).filter(models.Drawing.project_id == project_id)
    if drawing_ids:
        try:
            id_filter = [int(x) for x in drawing_ids.split(",") if x.strip()]
        except ValueError:
            raise HTTPException(status_code=400, detail="drawing_ids must be a comma-separated list of integers")
        drawings_query = drawings_query.filter(models.Drawing.id.in_(id_filter))
    drawings = drawings_query.order_by(models.Drawing.page_number, models.Drawing.uploaded_at).all()

    all_rows = []
    for drawing in drawings:
        all_rows.extend(export_engine.extract_rows(db, drawing))
    available_trades = sorted({r["trade"] for r in all_rows})

    trade_filter = [t.strip() for t in trades.split(",") if t.strip()] if trades else None
    filtered_rows = export_engine.filter_rows(all_rows, trades=trade_filter)
    final_rows = export_engine.apply_multiplier(filtered_rows, multiplier)

    return {
        "rows": final_rows,
        "available_trades": available_trades,
        "drawings": [
            {"id": d.id, "name": d.sheet_number or d.sheet_name or d.original_filename}
            for d in drawings
        ],
    }


class GenerateExportRow(BaseModel):
    row_id: str
    drawing_id: int
    drawing_name: str
    trade: str
    item: str
    quantity: float
    unit: str


class GenerateExportRequest(BaseModel):
    format: str  # 'excel' | 'csv' | 'pdf'
    rows: List[GenerateExportRow]
    group_by: Optional[List[str]] = None  # up to 3, from export_engine.GROUP_DIMENSIONS
    title: Optional[str] = None


_EXPORT_MEDIA_TYPES = {
    "excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "csv": "text/csv",
    "pdf": "application/pdf",
}
_EXPORT_EXTENSIONS = {"excel": "xlsx", "csv": "csv", "pdf": "pdf"}


@router.post("/projects/{project_id}/generate")
async def generate_project_export(
    project_id: int,
    payload: GenerateExportRequest,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Renders exactly the rows it's given, verbatim — this is what makes the
    inline editable grid real rather than cosmetic: whatever the frontend
    submits here (after a user edited quantities or excluded rows in the
    preview) is what ends up in the file, with no re-query in between that
    could silently diverge from what was shown/edited.
    """
    project = _get_project_for_export(project_id, current_user, db)

    if payload.format not in _EXPORT_MEDIA_TYPES:
        raise HTTPException(status_code=400, detail="format must be one of: excel, csv, pdf")
    if not payload.rows:
        raise HTTPException(status_code=400, detail="No rows to export")

    rows = [r.model_dump() for r in payload.rows]
    title = payload.title or f"{project.name} — Takeoff Export"

    try:
        file_content = export_engine.generate_report(rows, payload.format, group_by=payload.group_by, title=title)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    safe_name = "".join(c if c.isalnum() or c in "-_" else "_" for c in project.name) or "export"
    filename = f"takeoff_{safe_name}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.{_EXPORT_EXTENSIONS[payload.format]}"

    return StreamingResponse(
        file_content,
        media_type=_EXPORT_MEDIA_TYPES[payload.format],
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def generate_excel_export(drawing_data, result_data):
    """Generate Excel file from takeoff data"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Takeoff Report"
    
    # Styling
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title section
    ws['A1'] = "TakeOff.ai - Takeoff Report"
    ws['A1'].font = Font(bold=True, size=16, color="1F2937")
    ws.merge_cells('A1:F1')
    
    # Drawing info
    ws['A3'] = "Drawing Information"
    ws['A3'].font = Font(bold=True, size=14)
    
    ws['A4'] = "File Name:"
    ws['B4'] = drawing_data.get('filename', 'N/A')
    ws['A5'] = "Sheet Name:"
    ws['B5'] = drawing_data.get('sheet_name', 'N/A')
    ws['A6'] = "Upload Date:"
    ws['B6'] = drawing_data.get('uploaded_at', 'N/A')
    ws['A7'] = "Processing Status:"
    ws['B7'] = drawing_data.get('processing_status', 'N/A')
    
    # Parse detection data
    try:
        detection = json.loads(result_data.get('detection_data', '{}'))
    except:
        detection = {}
    
    # Quantities section
    start_row = 10
    ws[f'A{start_row}'] = "Quantities Breakdown"
    ws[f'A{start_row}'].font = Font(bold=True, size=14)
    
    # Headers
    header_row = start_row + 2
    headers = ['Item', 'Trade', 'Quantity', 'Unit', 'Notes']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Populate quantities
    quantities = detection.get('quantities', [])
    current_row = header_row + 1
    for item in quantities:
        ws.cell(row=current_row, column=1, value=item.get('item', 'N/A')).border = border
        ws.cell(row=current_row, column=2, value=item.get('trade', 'N/A')).border = border
        ws.cell(row=current_row, column=3, value=item.get('quantity', 0)).border = border
        ws.cell(row=current_row, column=4, value=item.get('unit', 'N/A')).border = border
        ws.cell(row=current_row, column=5, value='').border = border
        current_row += 1
    
    # Detections summary
    summary_row = current_row + 2
    ws[f'A{summary_row}'] = "Detection Summary"
    ws[f'A{summary_row}'].font = Font(bold=True, size=14)
    
    summary = detection.get('summary', {})
    summary_data = [
        ('Total Rooms', summary.get('rooms', 0)),
        ('Total Doors', summary.get('doors', 0)),
        ('Total Windows', summary.get('windows', 0)),
        ('Total Area (SF)', summary.get('totalArea', 0)),
    ]
    
    summary_start = summary_row + 2
    for idx, (label, value) in enumerate(summary_data):
        ws.cell(row=summary_start + idx, column=1, value=label).font = Font(bold=True)
        ws.cell(row=summary_start + idx, column=2, value=value)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 15
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def generate_csv_export(drawing_data, result_data):
    """Generate CSV file from takeoff data"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header section
    writer.writerow(['TakeOff.ai - Takeoff Report'])
    writer.writerow([])
    
    # Drawing info
    writer.writerow(['Drawing Information'])
    writer.writerow(['File Name:', drawing_data.get('filename', 'N/A')])
    writer.writerow(['Sheet Name:', drawing_data.get('sheet_name', 'N/A')])
    writer.writerow(['Upload Date:', drawing_data.get('uploaded_at', 'N/A')])
    writer.writerow(['Processing Status:', drawing_data.get('processing_status', 'N/A')])
    writer.writerow([])
    
    # Parse detection data
    try:
        detection = json.loads(result_data.get('detection_data', '{}'))
    except:
        detection = {}
    
    # Quantities
    writer.writerow(['Quantities Breakdown'])
    writer.writerow(['Item', 'Trade', 'Quantity', 'Unit', 'Notes'])
    
    quantities = detection.get('quantities', [])
    for item in quantities:
        writer.writerow([
            item.get('item', 'N/A'),
            item.get('trade', 'N/A'),
            item.get('quantity', 0),
            item.get('unit', 'N/A'),
            ''
        ])
    
    writer.writerow([])
    
    # Detection summary
    writer.writerow(['Detection Summary'])
    summary = detection.get('summary', {})
    writer.writerow(['Total Rooms', summary.get('rooms', 0)])
    writer.writerow(['Total Doors', summary.get('doors', 0)])
    writer.writerow(['Total Windows', summary.get('windows', 0)])
    writer.writerow(['Total Area (SF)', summary.get('totalArea', 0)])
    
    # Convert to bytes
    output.seek(0)
    return io.BytesIO(output.getvalue().encode('utf-8'))

@router.get("/drawings/{drawing_id}/{format}")
async def export_drawing(
    drawing_id: int,
    format: str,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Export takeoff data for a specific drawing
    Formats: excel, csv
    """
    # Validate format
    if format not in ['excel', 'csv']:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid format. Must be 'excel' or 'csv'"
        )
    
    # Verify drawing access
    drawing = db.query(models.Drawing).join(models.Project).filter(
        models.Drawing.id == drawing_id,
        models.Project.organization_id == current_user.organization_id
    ).first()
    
    if not drawing:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Drawing not found"
        )
    
    # Get latest takeoff result
    result = db.query(models.TakeoffResult).filter(
        models.TakeoffResult.drawing_id == drawing_id
    ).order_by(models.TakeoffResult.created_at.desc()).first()
    
    if not result:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No takeoff results found for this drawing"
        )
    
    # Prepare data
    drawing_data = {
        'filename': drawing.original_filename,
        'sheet_name': drawing.sheet_name or 'N/A',
        'uploaded_at': drawing.uploaded_at.strftime('%Y-%m-%d %H:%M:%S') if drawing.uploaded_at else 'N/A',
        'processing_status': drawing.processing_status.value if drawing.processing_status else 'N/A'
    }
    
    result_data = {
        'detection_data': result.detection_data,
        'quantities_data': result.quantities_data
    }
    
    # Generate file
    if format == 'excel':
        file_content = generate_excel_export(drawing_data, result_data)
        filename = f"takeoff_{drawing.original_filename.rsplit('.', 1)[0]}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:  # csv
        file_content = generate_csv_export(drawing_data, result_data)
        filename = f"takeoff_{drawing.original_filename.rsplit('.', 1)[0]}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.csv"
        media_type = "text/csv"
    
    return StreamingResponse(
        file_content,
        media_type=media_type,
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )

@router.get("/projects/{project_id}/{format}")
async def export_project(
    project_id: int,
    format: str,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Export all takeoff data for a project
    Formats: excel, csv
    """
    # Validate format
    if format not in ['excel', 'csv']:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid format. Must be 'excel' or 'csv'"
        )
    
    # Verify project access
    project = db.query(models.Project).filter(
        models.Project.id == project_id,
        models.Project.organization_id == current_user.organization_id
    ).first()
    
    if not project:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found"
        )
    
    # Get all drawings with results
    drawings = db.query(models.Drawing).filter(
        models.Drawing.project_id == project_id
    ).all()
    
    if not drawings:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No drawings found for this project"
        )
    
    # For simplicity, export first drawing (can be extended to include all)
    # In production, you'd combine all drawings into one report
    first_drawing = drawings[0]
    result = db.query(models.TakeoffResult).filter(
        models.TakeoffResult.drawing_id == first_drawing.id
    ).order_by(models.TakeoffResult.created_at.desc()).first()
    
    if not result:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No takeoff results found for this project"
        )
    
    # Prepare data
    drawing_data = {
        'filename': f"{project.name} - {first_drawing.original_filename}",
        'sheet_name': first_drawing.sheet_name or 'N/A',
        'uploaded_at': first_drawing.uploaded_at.strftime('%Y-%m-%d %H:%M:%S') if first_drawing.uploaded_at else 'N/A',
        'processing_status': first_drawing.processing_status.value if first_drawing.processing_status else 'N/A'
    }
    
    result_data = {
        'detection_data': result.detection_data,
        'quantities_data': result.quantities_data
    }
    
    # Generate file
    if format == 'excel':
        file_content = generate_excel_export(drawing_data, result_data)
        filename = f"project_{project.name}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:  # csv
        file_content = generate_csv_export(drawing_data, result_data)
        filename = f"project_{project.name}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.csv"
        media_type = "text/csv"
    
    return StreamingResponse(
        file_content,
        media_type=media_type,
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )



