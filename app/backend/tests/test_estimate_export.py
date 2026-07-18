"""Tests for saved-estimate Excel export + serialization (pure — no DB)."""

import json
from io import BytesIO
from types import SimpleNamespace

from estimating.estimate_export import estimate_to_excel
from estimating.persistence import estimate_to_dict

SNAPSHOT = {
    "drivers": {"floor_area_sf": 1800, "wall_lf": 240, "door_count": 14},
    "line_items": [
        {"item": "Flooring material", "trade": "Flooring", "quantity": 1980.0, "unit": "sf",
         "unit_cost": 3.5, "amount": 6930.0},
        {"item": "Gypsum board 5/8\"", "trade": "Drywall", "quantity": 4752.0, "unit": "sf",
         "unit_cost": 0.55, "amount": 2613.6},
    ],
    "by_trade": {"Flooring": 6930.0, "Drywall": 2613.6},
    "total": 9543.6,
}


def test_estimate_to_excel_returns_valid_xlsx():
    data = estimate_to_excel(SNAPSHOT, title="My Estimate")
    assert isinstance(data, bytes) and len(data) > 0
    assert data[:2] == b"PK"        # xlsx is a zip container

    # Round-trip: openpyxl can read it back and the total is present.
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(data))
    ws = wb.active
    flat = [c.value for row in ws.iter_rows() for c in row]
    assert "My Estimate" in flat
    assert "Flooring material" in flat
    assert 9543.6 in flat           # grand total rendered


def test_estimate_to_excel_handles_empty():
    data = estimate_to_excel({"line_items": [], "by_trade": {}, "total": 0})
    assert data[:2] == b"PK"


def test_estimate_to_dict_parses_snapshot():
    est = SimpleNamespace(
        id=5, name="E1", project_id=2, drawing_id=9, cost_book_id=3,
        total=9543.6, data=json.dumps(SNAPSHOT), created_at=None,
    )
    d = estimate_to_dict(est)
    assert d["id"] == 5 and d["total"] == 9543.6
    assert d["data"]["by_trade"]["Flooring"] == 6930.0


def test_estimate_to_dict_list_view_omits_data():
    est = SimpleNamespace(id=5, name="E1", project_id=2, drawing_id=9, cost_book_id=None,
                          total=1.0, data="{}", created_at=None)
    d = estimate_to_dict(est, include_data=False)
    assert "data" not in d and d["name"] == "E1"


def test_estimate_to_dict_tolerates_bad_json():
    est = SimpleNamespace(id=1, name="x", project_id=None, drawing_id=None, cost_book_id=None,
                          total=0, data="not json", created_at=None)
    assert estimate_to_dict(est)["data"] == {}
