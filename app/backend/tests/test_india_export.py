"""
Tests for BOQ Excel/PDF export (increment #5). Renders a real estimate to bytes
and validates the output is a well-formed workbook / PDF containing BOQ data.
"""

import io
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from estimating import RateBook, full_estimate, boq_to_excel, boq_to_pdf

_MEASURE = {"summary": {"totalArea": 1000.0, "walls_lf": 200.0, "rooms": 4}}


def _estimate():
    est = full_estimate(_MEASURE, RateBook.sample())
    est["rate_book_note"] = "sample"
    return est


def test_excel_export_is_valid_workbook_with_boq():
    from openpyxl import load_workbook

    data = boq_to_excel(_estimate())
    assert isinstance(data, bytes) and len(data) > 0
    wb = load_workbook(io.BytesIO(data))
    ws = wb["BOQ"]
    text = "\n".join(
        str(c.value) for row in ws.iter_rows() for c in row if c.value is not None
    )
    assert "Bill of Quantities" in text
    assert "Grand total" in text
    # A priced flooring line from the sample book is present.
    assert "Vitrified" in text


def test_pdf_export_is_valid_pdf():
    data = boq_to_pdf(_estimate())
    assert isinstance(data, bytes) and len(data) > 0
    assert data[:5] == b"%PDF-"


def test_pdf_export_handles_empty_boq():
    est = full_estimate({"summary": {}}, RateBook.sample())
    data = boq_to_pdf(est)
    assert data[:5] == b"%PDF-"


def test_excel_export_handles_empty_boq():
    from openpyxl import load_workbook

    est = full_estimate({"summary": {}}, RateBook.sample())
    data = boq_to_excel(est)
    wb = load_workbook(io.BytesIO(data))
    assert "BOQ" in wb.sheetnames
