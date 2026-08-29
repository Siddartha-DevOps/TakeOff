import csv
import io
import inspect
import json
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from canonical_takeoff import (
    CanonicalAnnotationError,
    build_quantities,
    geometry_element,
    normalize_annotations,
    synchronize_corrected_takeoff,
    validate_conditions,
)
import models
from export_engine import extract_rows
from routes.export_routes import generate_csv_export, generate_excel_export
from routes.assemblies_routes import _compute_drawing_estimate
from routes.takeoff_routes import restore_annotation_history


SCALE_ONE_FOOT_PER_UNIT = 72 * 12


def annotation(annotation_id, annotation_type, geometry, *, label, layer="manual", holes=None, meta=None):
    payload = {
        "id": annotation_id,
        "type": annotation_type,
        "geometry": geometry,
        "style": {},
        "layerId": layer,
        "source": "manual",
        "meta": {"label": label, **(meta or {})},
        "measuredValue": 999999,  # prove the server never trusts the browser value
    }
    if holes is not None:
        payload["holes"] = holes
    return payload


def canonical(items):
    return normalize_annotations(items, SCALE_ONE_FOOT_PER_UNIT, "PDF")


def quantity_total(items, unit):
    return sum(row["quantity"] for row in build_quantities(canonical(items)) if row["unit"] == unit)


def test_edit_room_polygon_updates_area():
    room = annotation("room-1", "area", [[0, 0], [10, 0], [10, 10], [0, 10]], label="Office", layer="rooms")
    assert quantity_total([room], "sf") == 100

    room["geometry"] = [[0, 0], [20, 0], [20, 10], [0, 10]]
    assert quantity_total([room], "sf") == 200


def test_delete_annotation_removes_quantity():
    room = annotation("room-1", "area", [[0, 0], [10, 0], [10, 10], [0, 10]], label="Office", layer="rooms")
    assert quantity_total([room], "sf") == 100
    assert quantity_total([], "sf") == 0


def test_split_annotation_recalculates_both_children():
    original = annotation("room-1", "area", [[0, 0], [10, 0], [10, 10], [0, 10]], label="Office", layer="rooms")
    left = annotation("room-1-a", "area", [[0, 0], [4, 0], [4, 10], [0, 10]], label="Office", layer="rooms")
    right = annotation("room-1-b", "area", [[4, 0], [10, 0], [10, 10], [4, 10]], label="Office", layer="rooms")
    assert quantity_total([original], "sf") == 100
    split = canonical([left, right])
    assert [item["measuredValue"] for item in split] == [40, 60]
    assert sum(row["quantity"] for row in build_quantities(split)) == 100


def test_merge_annotations_recalculates_merged_polygon():
    left = annotation("a", "area", [[0, 0], [4, 0], [4, 10], [0, 10]], label="Office", layer="rooms")
    right = annotation("b", "area", [[4, 0], [10, 0], [10, 10], [4, 10]], label="Office", layer="rooms")
    merged = annotation("merged", "area", [[0, 0], [10, 0], [10, 10], [0, 10]], label="Office", layer="rooms")
    assert quantity_total([left, right], "sf") == 100
    normalized = canonical([merged])
    assert len(normalized) == 1
    assert normalized[0]["measuredValue"] == 100


def test_add_and_remove_hole_updates_net_area_and_postgis_polygon():
    outer = [[0, 0], [10, 0], [10, 10], [0, 10]]
    hole = [[2, 2], [6, 2], [6, 7], [2, 7]]
    with_hole = canonical([annotation("room", "area", outer, holes=[hole], label="Office", layer="rooms")])[0]
    without_hole = canonical([annotation("room", "area", outer, holes=[], label="Office", layer="rooms")])[0]
    assert with_hole["measuredValue"] == 80
    assert without_hole["measuredValue"] == 100
    assert str(geometry_element(with_hole)).count("(") >= 3  # polygon plus an interior ring


def test_wall_linear_edit_and_arc_are_measured_in_lf():
    wall = annotation("wall", "line", [[0, 0], [3, 4]], label="Interior wall", layer="walls")
    assert quantity_total([wall], "lf") == 5
    wall["geometry"] = [[0, 0], [6, 8]]
    assert quantity_total([wall], "lf") == 10

    arc = annotation("arc", "line", [[1, 0], [0, 1], [-1, 0]], label="Curved wall", layer="walls", meta={"curve": "arc"})
    assert canonical([arc])[0]["measuredValue"] == pytest.approx(3.1416)


def test_count_edit_stays_one_and_delete_removes_count():
    symbol = annotation("door", "count", [[0, 0]], label="Door", layer="doors")
    assert quantity_total([symbol], "ea") == 1
    symbol["geometry"] = [[500, 900]]
    assert quantity_total([symbol], "ea") == 1
    assert quantity_total([], "ea") == 0


def test_save_reload_preserves_server_corrected_quantities():
    submitted = [annotation("room", "area", [[0, 0], [12, 0], [12, 8], [0, 8]], label="Bedroom", layer="rooms")]
    saved = canonical(submitted)
    reloaded = normalize_annotations(json.loads(json.dumps(saved)), SCALE_ONE_FOOT_PER_UNIT, "PDF")
    assert reloaded == saved
    assert build_quantities(reloaded) == build_quantities(saved)
    assert reloaded[0]["measuredValue"] == 96


class NoQueryDB:
    def query(self, *_args, **_kwargs):
        raise AssertionError("annotation-backed exports must not query a stale TakeoffResult")


def test_corrected_quantities_appear_in_csv_and_excel_exports():
    saved = canonical([
        annotation("room", "area", [[0, 0], [12, 0], [12, 8], [0, 8]], label="Bedroom", layer="rooms"),
        annotation("door", "count", [[3, 3]], label="Door", layer="doors"),
    ])
    drawing = SimpleNamespace(
        id=7, project_id=3, annotations_data=json.dumps(saved),
        scale_ratio=SCALE_ONE_FOOT_PER_UNIT, file_type="PDF",
        sheet_number="A-101", sheet_name=None, original_filename="plan.pdf",
    )
    rows = extract_rows(NoQueryDB(), drawing)
    assert {(row["item"], row["quantity"], row["unit"]) for row in rows} == {
        ("Bedroom area", 96, "sf"), ("Door", 1, "ea")
    }

    result_data = {"detection_data": json.dumps({"summary": {}}), "quantities_data": json.dumps([
        {"item": row["item"], "trade": row["trade"], "quantity": row["quantity"], "unit": row["unit"]}
        for row in rows
    ])}
    csv_bytes = generate_csv_export({"filename": "plan.pdf"}, result_data).getvalue().decode("utf-8")
    csv_rows = list(csv.reader(io.StringIO(csv_bytes)))
    assert ["Bedroom area", "Flooring", "96.0", "sf", ""] in csv_rows
    assert ["Door", "Doors", "1.0", "ea", ""] in csv_rows

    excel_bytes = generate_excel_export({"filename": "plan.pdf"}, result_data).getvalue()
    worksheet = load_workbook(io.BytesIO(excel_bytes), read_only=True).active
    values = [tuple(row) for row in worksheet.iter_rows(values_only=True)]
    assert any(row[:5] == ("Bedroom area", "Flooring", 96, "sf", None) for row in values)
    assert any(row[:5] == ("Door", "Doors", 1, "ea", None) for row in values)


def test_condition_assignment_is_project_scoped_and_type_safe():
    area = canonical([annotation("room", "area", [[0, 0], [5, 0], [5, 5], [0, 5]], label="Office", meta={"conditionId": 99})])[0]
    with pytest.raises(CanonicalAnnotationError, match="outside this project"):
        validate_conditions([area], {})

    wrong_type = SimpleNamespace(id=99, annotation_type="line", unit="lf", trade="Framing", name="Wall")
    with pytest.raises(CanonicalAnnotationError, match="not compatible"):
        validate_conditions([area], {99: wrong_type})

    valid = SimpleNamespace(id=99, annotation_type="area", unit="sf", trade="Flooring", name="Carpet")
    validate_conditions([area], {99: valid})
    assert build_quantities([area], {99: valid}) == [
        {"trade": "Flooring", "item": "Carpet", "quantity": 25.0, "unit": "sf"}
    ]


def test_rejected_annotation_remains_versionable_but_has_no_downstream_quantity():
    rejected = canonical([annotation(
        "room", "area", [[0, 0], [10, 0], [10, 10], [0, 10]],
        label="False positive", layer="rooms", meta={"rejected": True},
    )])
    assert len(rejected) == 1
    assert build_quantities(rejected) == []


class FakeQuery:
    def __init__(self, rows):
        self.rows = rows

    def filter(self, *_args):
        return self

    def join(self, *_args):
        return self

    def order_by(self, *_args):
        return self

    def all(self):
        return list(self.rows)

    def first(self):
        return self.rows[0] if self.rows else None

    def delete(self, **_kwargs):
        count = len(self.rows)
        self.rows.clear()
        return count


class ProjectionDB:
    def __init__(self, *, result=None, conditions=None, old_detections=None, embeddings=None, drawings=None):
        self.result_rows = [result] if result else []
        self.condition_rows = list(conditions or [])
        self.detection_rows = list(old_detections or [])
        self.embedding_rows = list(embeddings or [])
        self.drawing_rows = list(drawings or [])
        self.added = []
        self.deleted = []
        self.next_detection_id = 100

    def query(self, model):
        return FakeQuery({
            models.Condition: self.condition_rows,
            models.Detection: self.detection_rows,
            models.DrawingEmbedding: self.embedding_rows,
            models.TakeoffResult: self.result_rows,
            models.Drawing: self.drawing_rows,
        }[model])

    def add(self, value):
        if isinstance(value, models.Detection) and value.id is None:
            value.id = self.next_detection_id
            self.next_detection_id += 1
        self.added.append(value)

    def delete(self, value):
        self.deleted.append(value)

    def flush(self):
        pass


def test_sync_refreshes_postgis_measurements_conditions_results_and_search_rows():
    result = SimpleNamespace(detection_data="old", quantities_data="old")
    old_detection = SimpleNamespace(id=1)
    db = ProjectionDB(result=result, old_detections=[old_detection], embeddings=[SimpleNamespace(id=2)])
    drawing = SimpleNamespace(
        id=7, project_id=3, scale_ratio=SCALE_ONE_FOOT_PER_UNIT,
        file_type="PDF", annotations_data=None,
    )
    payload = [
        annotation("room", "area", [[0, 0], [10, 0], [10, 8], [0, 8]], label="Office", layer="rooms"),
        annotation("wall", "line", [[0, 0], [10, 0]], label="Wall", layer="walls"),
        annotation("door", "count", [[2, 2]], label="Door", layer="doors"),
    ]

    projection = synchronize_corrected_takeoff(db, drawing, payload)
    detections = [value for value in db.added if isinstance(value, models.Detection)]
    measurements = [value for value in db.added if isinstance(value, models.Measurement)]
    assert old_detection in db.deleted
    assert db.embedding_rows == []
    assert len(detections) == len(measurements) == 3
    assert {(row.annotation_type, row.condition_id) for row in detections} == {
        ("area", None), ("line", None), ("count", None)
    }
    assert {(row.value, row.unit) for row in measurements} == {(80, "sf"), (10, "lf"), (1, "ea")}
    assert json.loads(result.quantities_data) == projection["quantities"]
    assert json.loads(result.detection_data)["summary"] == projection["detection"]["summary"]


def test_history_restore_routes_through_the_same_canonical_synchronizer():
    source = inspect.getsource(restore_annotation_history)
    assert "synchronize_corrected_takeoff" in source
    assert 'projection["quantities"]' in source
    assert "models.Project.organization_id == current_user.organization_id" in source
    assert "models.AnnotationRevision.drawing_id == drawing_id" in source


def test_history_snapshot_restore_rebuilds_the_prior_quantity_projection():
    result = SimpleNamespace(detection_data="old", quantities_data="old")
    db = ProjectionDB(result=result)
    drawing = SimpleNamespace(
        id=7, project_id=3, scale_ratio=SCALE_ONE_FOOT_PER_UNIT,
        file_type="PDF", annotations_data=None,
    )
    prior_snapshot = [annotation(
        "room", "area", [[0, 0], [10, 0], [10, 10], [0, 10]],
        label="Office", layer="rooms",
    )]
    edited_snapshot = [annotation(
        "room", "area", [[0, 0], [20, 0], [20, 10], [0, 10]],
        label="Office", layer="rooms",
    )]
    synchronize_corrected_takeoff(db, drawing, edited_snapshot)
    assert json.loads(result.quantities_data)[0]["quantity"] == 200

    restored = synchronize_corrected_takeoff(db, drawing, json.loads(json.dumps(prior_snapshot)))
    assert restored["quantities"][0]["quantity"] == 100
    assert json.loads(result.quantities_data) == restored["quantities"]


def test_estimating_uses_corrected_annotation_quantities_not_stale_result():
    corrected = canonical([
        annotation("floor", "area", [[0, 0], [20, 0], [20, 10], [0, 10]], label="Floor", layer="rooms"),
        annotation("wall", "line", [[0, 0], [25, 0]], label="Interior wall", layer="walls"),
        annotation("door", "count", [[5, 5]], label="Door", layer="doors"),
    ])
    drawing = SimpleNamespace(
        id=44, project_id=3, annotations_data=json.dumps(corrected),
        scale_ratio=SCALE_ONE_FOOT_PER_UNIT, file_type="PDF",
    )
    stale_result = SimpleNamespace(quantities_data=json.dumps([
        {"trade": "Flooring", "item": "Floor", "quantity": 9999, "unit": "sf"}
    ]))
    db = ProjectionDB(result=stale_result, drawings=[drawing])
    estimate = _compute_drawing_estimate(44, None, SimpleNamespace(organization_id=8), db)
    assert estimate["drivers"] == {"floor_area_sf": 200.0, "wall_lf": 25.0, "door_count": 1.0}
